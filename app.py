"""
Helmet Violation Detection System
"""

import streamlit as st
import cv2
import numpy as np
import re
import io
import torch
import base64
from datetime import datetime
from PIL import Image
from pathlib import Path

st.set_page_config(
    page_title="Helmet Violation Detection",
    page_icon="🪖",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
* { font-family: 'DM Sans', sans-serif !important; }

.sc { display:flex; gap:12px; margin:12px 0 20px; }
.sc-item { flex:1; padding:18px 12px; border-radius:10px; text-align:center;
    border:1px solid #e2e8f0; background:#fff; }
.sc-num { font-size:2rem; font-weight:700;
    font-family:'DM Mono',monospace !important; line-height:1; }
.sc-lbl { font-size:0.65rem; font-weight:600; text-transform:uppercase;
    letter-spacing:0.1em; color:#94a3b8; margin-top:6px; }
.sc-blue  .sc-num { color:#1d4ed8; }
.sc-green .sc-num { color:#16a34a; }
.sc-red   .sc-num { color:#dc2626; }

.plate-ok { background:#fefce8; border:2px solid #ca8a04; border-radius:8px;
    padding:14px 20px; font-family:'DM Mono',monospace !important; font-size:1.5rem;
    font-weight:600; color:#713f12; text-align:center; letter-spacing:0.2em; margin:8px 0; }
.plate-na { background:#f8fafc; border:1.5px dashed #cbd5e1; border-radius:8px;
    padding:12px 20px; font-size:0.85rem; color:#94a3b8; text-align:center; margin:8px 0; }

.vtag { display:inline-flex; align-items:center; gap:6px; background:#fef2f2;
    border:1px solid #fca5a5; color:#b91c1c; border-radius:20px;
    padding:4px 14px; font-size:0.78rem; font-weight:600; margin:3px; }

.sec-label { font-size:0.65rem; font-weight:700; letter-spacing:0.12em;
    text-transform:uppercase; color:#94a3b8; margin:16px 0 6px; }

.empty-state { border:2px dashed #e2e8f0; border-radius:12px; padding:64px 20px;
    text-align:center; background:#fafbfc; }
</style>
""", unsafe_allow_html=True)

# ── Config ────────────────────────────────────────────────────────────────────
MODEL_DIR         = Path("models")
MODEL_DIR.mkdir(exist_ok=True)
HELMET_MODEL_PATH = MODEL_DIR / "helmet_best.pt"
PLATE_MODEL_PATH  = MODEL_DIR / "plate_best.pt"
GEMINI_API_KEY    = ""


# ── Load models ───────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def load_models():
    from ultralytics import YOLO
    from fast_alpr import ALPR
    if not HELMET_MODEL_PATH.exists():
        return None
    m = {}
    m['helmet'] = YOLO(str(HELMET_MODEL_PATH))
    m['plate']  = YOLO(str(PLATE_MODEL_PATH)) if PLATE_MODEL_PATH.exists() else None
    m['person'] = YOLO('yolov8m.pt')
    m['alpr']   = ALPR(
        detector_model='yolo-v9-t-384-license-plate-end2end',
        ocr_model='cct-s-v2-global-model',
        detector_conf_thresh=0.10,
    )
    return m


# ── Helpers ───────────────────────────────────────────────────────────────────
def classify(label, idx):
    l = label.lower().replace(' ', '_')
    if any(k in l for k in ['without', 'no_helmet', 'nohelmet']): return 'no_helmet'
    if any(k in l for k in ['with_helmet', 'withhelmet', 'helmet']): return 'helmet'
    return 'no_helmet' if idx == 0 else 'helmet'

def iou(a, b):
    ax1,ay1,ax2,ay2=a; bx1,by1,bx2,by2=b
    ix1=max(ax1,bx1); iy1=max(ay1,by1); ix2=min(ax2,bx2); iy2=min(ay2,by2)
    inter=max(0,ix2-ix1)*max(0,iy2-iy1)
    if inter==0: return 0.0
    return inter/((ax2-ax1)*(ay2-ay1)+(bx2-bx1)*(by2-by1)-inter)

def draw_circle(f, x1, y1, x2, y2, tag, color):
    cx=int((x1+x2)/2); cy=int((y1+y2)/2)
    r=int(max(x2-x1,y2-y1)/2)+8
    cv2.circle(f,(cx,cy),r,color,2)
    (tw,th),_=cv2.getTextSize(tag,cv2.FONT_HERSHEY_SIMPLEX,0.6,2)
    cv2.putText(f,tag,(cx-tw//2,cy-r-6),cv2.FONT_HERSHEY_SIMPLEX,0.6,color,2)


# ── OCR  ────────────────────────────────────────────────────────────
def gemini_ocr(img):
    """Send image to Gemini, get plate number back."""
    import requests
    _, buf = cv2.imencode('.jpg', img)
    b64    = base64.b64encode(buf).decode('utf-8')
    try:
        resp = requests.post(
            f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}",
            json={"contents": [{"parts": [
                {"text": ("Find and read the vehicle number plate in this image. "
                          "Return ONLY the plate number text, nothing else. "
                          "If no plate is visible, return NONE.")},
                {"inline_data": {"mime_type": "image/jpeg", "data": b64}}
            ]}]},
            timeout=20
        )
        if resp.status_code != 200:
            return None
        text  = resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip().upper()
        if text in ("NONE", ""):
            return None
        text  = re.sub(r"[^A-Z0-9 ]", " ", text)
        text  = re.sub(r"\s+", " ", text).strip()
        clean = re.sub(r"[^A-Z0-9]", "", text)
        if len(clean) >= 4 and re.search(r"[A-Z]", clean) and re.search(r"[0-9]", clean):
            return text
    except Exception:
        pass
    return None


# ── Plate detection ───────────────────────────────────────────────────────────
def find_plate(img, rf, m):
    """
    1. ALPR + YOLO  → find plate bounding box → draw ellipse + save crop
    2. Gemini        → read plate text from FULL image (most reliable)
    Returns (plate_text, plate_crop_bgr)
    """
    H, W  = img.shape[:2]
    plate_crop = None

    # ── Step 1: Locate plate bbox ──────────────────────────────────────────
    candidates = []
    for p in m['alpr'].predict(img):
        b = p.detection.bounding_box
        x1,y1,x2,y2 = int(b.x1),int(b.y1),int(b.x2),int(b.y2)
        candidates.append(((x2-x1)*(y2-y1), x1,y1,x2,y2))

    if m['plate']:
        for box in m['plate'].predict(img, conf=0.10, verbose=False)[0].boxes:
            x1,y1,x2,y2 = map(int, box.xyxy[0].tolist())
            candidates.append(((x2-x1)*(y2-y1), x1,y1,x2,y2))

    if candidates:
        # Pick largest bbox — most likely the actual plate
        candidates.sort(reverse=True)
        _, x1,y1,x2,y2 = candidates[0]
        pad = 8
        px1,py1 = max(0,x1-pad), max(0,y1-pad)
        px2,py2 = min(W,x2+pad), min(H,y2+pad)
        plate_crop = img[py1:py2, px1:px2].copy()

        # Draw ellipse on result frame
        cv2.ellipse(rf,
            ((px1+px2)//2, (py1+py2)//2),
            ((px2-px1)//2+10, (py2-py1)//2+10),
            0, 0, 360, (0,200,255), 3)

    # ── Step 2: Gemini reads text from full image ──────────────────────────
    text = gemini_ocr(img)
    if text:
        return text, plate_crop

    return 'NOT DETECTED', plate_crop


# ── Helmet detection ──────────────────────────────────────────────────────────
def run_detection(frame, m, conf):
    H,W   = frame.shape[:2]
    result = frame.copy()
    counts = {'helmet':0,'no_helmet':0}
    violations = []
    seen = set()
    C_OK  = (0,180,0)
    C_BAD = (0,0,210)

    hboxes = []
    for box in m['helmet'].predict(frame,conf=conf,iou=0.45,verbose=False)[0].boxes:
        cv_val = float(box.conf[0]); idx = int(box.cls[0])
        status = classify(m['helmet'].names[idx], idx)
        x1,y1,x2,y2 = map(int, box.xyxy[0].tolist())
        key = (int((x1+x2)/2)//20, int((y1+y2)/2)//20)
        if key in seen: continue
        seen.add(key)
        hboxes.append({'status':status,'bbox':[x1,y1,x2,y2],'conf':cv_val})

    pboxes = []
    for box in m['person'].predict(frame,conf=0.25,iou=0.45,classes=[0],verbose=False)[0].boxes:
        x1,y1,x2,y2 = map(int, box.xyxy[0].tolist())
        pboxes.append([x1,y1,x2,y2])

    matched = set()
    for hb in hboxes:
        x1,y1,x2,y2 = hb['bbox']; status = hb['status']; counts[status] += 1
        tag = f"HELMET ({hb['conf']:.0%})" if status=='helmet' else f"NO HELMET ({hb['conf']:.0%})"
        draw_circle(result,x1,y1,x2,y2,tag,C_OK if status=='helmet' else C_BAD)
        if status=='no_helmet': violations.append({'bbox':[x1,y1,x2,y2]})
        for i,pb in enumerate(pboxes):
            px1,py1,px2,py2 = pb
            upper = [px1,py1,px2,py1+int((py2-py1)*0.4)]
            if iou([x1,y1,x2,y2],upper) > 0.05: matched.add(i)

    for i,pb in enumerate(pboxes):
        if i in matched: continue
        px1,py1,px2,py2 = pb
        if (px2-px1)*(py2-py1) < W*H*0.005: continue
        counts['no_helmet'] += 1
        draw_circle(result,px1,py1,px2,py2,'NO HELMET',C_BAD)
        violations.append({'bbox':[px1,py1,px2,py2]})

    if violations:
        plate_text, plate_crop = find_plate(frame, result, m)
    else:
        plate_text, plate_crop = 'N/A', None

    return result, counts, violations, plate_text, plate_crop


# ── Excel report ──────────────────────────────────────────────────────────────
def build_excel(image_name, counts, violations, plate_text, plate_crop):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook(); ws = wb.active; ws.title = "Violation Report"

    hdr_fill    = PatternFill("solid", fgColor="1A1D2E")
    red_fill    = PatternFill("solid", fgColor="FEF2F2")
    green_fill  = PatternFill("solid", fgColor="F0FDF4")
    yellow_fill = PatternFill("solid", fgColor="FFFBEB")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    sub_fill    = PatternFill("solid", fgColor="E5E9F0")
    border = Border(
        left=Side(style='thin',color='E5E9F0'), right=Side(style='thin',color='E5E9F0'),
        top=Side(style='thin',color='E5E9F0'),  bottom=Side(style='thin',color='E5E9F0'),
    )

    def cell(ref, text, bold=False, color="1A1D2E", size=10, fill=white_fill, align="left"):
        ws[ref] = text
        ws[ref].font      = Font(name="Arial", bold=bold, color=color, size=size)
        ws[ref].fill      = fill
        ws[ref].border    = border
        ws[ref].alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    ws.merge_cells("A1:F1")
    cell("A1","HELMET VIOLATION DETECTION REPORT",bold=True,color="FFFFFF",size=14,fill=hdr_fill,align="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    ts = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    cell("A2",f"Image: {image_name}   |   Generated: {ts}",
         color="64748B",size=9,fill=PatternFill("solid",fgColor="F8F9FC"),align="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A4:F4")
    cell("A4","DETECTION SUMMARY",bold=True,size=10,fill=sub_fill)
    ws.row_dimensions[4].height = 22

    for c,t in [("A5","Metric"),("B5","Value"),("C5","Status")]:
        cell(c,t,bold=True,color="FFFFFF",fill=hdr_fill,align="center")
    ws.row_dimensions[5].height = 20

    total = counts['helmet'] + counts['no_helmet']
    for i,(lbl,val,status,fill,scol) in enumerate([
        ("Total Riders",str(total),"—",white_fill,"1A1D2E"),
        ("With Helmet",str(counts['helmet']),"COMPLIANT",green_fill,"166534"),
        ("No Helmet",str(counts['no_helmet']),"VIOLATION",red_fill,"991B1B"),
    ], start=6):
        ws.row_dimensions[i].height = 20
        cell(f"A{i}",lbl,fill=fill)
        cell(f"B{i}",val,bold=True,fill=fill,align="center")
        cell(f"C{i}",status,bold=True,color=scol,fill=fill,align="center")

    ws.merge_cells("A9:F9")
    cell("A9","VIOLATION DETAILS  (No Helmet Riders Only)",
         bold=True,color="991B1B",size=10,fill=PatternFill("solid",fgColor="FEF2F2"))
    ws.row_dimensions[9].height = 22

    for c,t in [("A10","Rider #"),("B10","Violation"),("C10","Number Plate"),("D10","Bounding Box")]:
        cell(c,t,bold=True,color="FFFFFF",fill=hdr_fill,align="center")
    ws.row_dimensions[10].height = 20

    if violations:
        for i,v in enumerate(violations,start=1):
            r = 10+i; ws.row_dimensions[r].height = 20
            bbox = f"x1={v['bbox'][0]} y1={v['bbox'][1]} x2={v['bbox'][2]} y2={v['bbox'][3]}"
            pval = plate_text if plate_text not in ('NOT DETECTED','N/A') else "—"
            cell(f"A{r}",i,bold=True,fill=red_fill,align="center")
            cell(f"B{r}","No Helmet",bold=True,color="991B1B",fill=red_fill)
            cell(f"C{r}",pval,bold=True,color="92400E",fill=yellow_fill,align="center")
            cell(f"D{r}",bbox,fill=red_fill)
    else:
        ws.merge_cells("A11:D11")
        cell("A11","No violations detected.",color="166534",fill=green_fill,align="center")
        ws.row_dimensions[11].height = 20

    plate_row = 10 + max(len(violations),1) + 3

    if violations and plate_crop is not None:
        ws.merge_cells(f"A{plate_row}:F{plate_row}")
        cell(f"A{plate_row}","NUMBER PLATE  (Violation Evidence)",
             bold=True,color="92400E",size=10,fill=yellow_fill)
        ws.row_dimensions[plate_row].height = 22

        ocr_row = plate_row+1
        ws.merge_cells(f"A{ocr_row}:F{ocr_row}")
        cell(f"A{ocr_row}",f"OCR Read:   {plate_text}",
             bold=True,color="92400E",size=14,fill=yellow_fill,align="center")
        ws.row_dimensions[ocr_row].height = 32

        img_row = ocr_row+1
        pil = Image.fromarray(cv2.cvtColor(plate_crop, cv2.COLOR_BGR2RGB))
        if pil.width < 400:
            scale = 400/pil.width
            pil = pil.resize((int(pil.width*scale),int(pil.height*scale)),Image.LANCZOS)
        buf_img = io.BytesIO(); pil.save(buf_img,format='PNG'); buf_img.seek(0)
        xl = XLImage(buf_img); xl.anchor = f"A{img_row}"; ws.add_image(xl)
        for r in range(img_row, img_row+25): ws.row_dimensions[r].height = 15

    elif violations:
        ws.merge_cells(f"A{plate_row}:F{plate_row}")
        cell(f"A{plate_row}",f"Number Plate (OCR):   {plate_text}",
             bold=True,color="92400E",size=13,fill=yellow_fill,align="center")
        ws.row_dimensions[plate_row].height = 32

    for col,w in [('A',10),('B',18),('C',20),('D',32),('E',12),('F',12)]:
        ws.column_dimensions[col].width = w

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🪖 Helmet Detector")
    st.divider()
    st.markdown("**Model Status**")
    st.success("helmet_best.pt ✓" if HELMET_MODEL_PATH.exists() else "helmet_best.pt ✗",
               icon="🟢" if HELMET_MODEL_PATH.exists() else "🔴")
    st.info("plate_best.pt ✓" if PLATE_MODEL_PATH.exists() else "plate_best.pt — ALPR fallback",
            icon="🟢" if PLATE_MODEL_PATH.exists() else "🟡")
    st.success("ALPR · Plate OCR ✓", icon="🟢")
    st.divider()
    st.markdown("**Settings**")
    conf_threshold = st.slider("Confidence", 0.20, 0.80, 0.45, 0.05,
                                help="Raise to reduce false positives")
    st.divider()
    st.markdown("**Steps**")
    st.markdown("1. Upload image\n2. Click **Run Detection**\n3. Download results")
    st.divider()
    st.caption("YOLOv8 · ALPR · Computer Vision")


# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("## 🪖 Helmet Violation Detection")
st.caption("Detect helmet violations and read number plates from motorcycle traffic images.")
st.divider()

if not HELMET_MODEL_PATH.exists():
    st.error(
        "**Helmet model not found.** Place `helmet_best.pt` in the `models/` folder.\n\n"
        "Download from Colab after training:\n"
        "```python\nfrom google.colab import files\n"
        "files.download('/content/runs/detect/helmet_v1/weights/best.pt')\n```"
    )
    st.stop()


# ── Layout ────────────────────────────────────────────────────────────────────
col_l, col_r = st.columns([1, 1], gap="large")

with col_l:
    st.markdown('<p class="sec-label">Input Image</p>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Upload image", type=["jpg","jpeg","png","webp"],
                                  label_visibility="collapsed")
    if uploaded:
        img_pil = Image.open(uploaded).convert("RGB")
        st.image(img_pil, use_container_width=True)
        st.markdown("")
        run_btn = st.button("▶  Run Detection", use_container_width=True)
    else:
        st.markdown("""
<div class="empty-state">
    <div style="font-size:2.6rem">📷</div>
    <div style="font-weight:600;color:#64748b;margin:8px 0 4px">Drop an image here</div>
    <div style="font-size:0.82rem;color:#94a3b8">JPG · PNG · WEBP</div>
</div>""", unsafe_allow_html=True)
        run_btn = False

with col_r:
    st.markdown('<p class="sec-label">Results</p>', unsafe_allow_html=True)

    if uploaded and run_btn:
        with st.spinner("Loading models…"):
            models = load_models()

        if not models:
            st.error("Failed to load models. Check that `helmet_best.pt` exists in `models/`.")
        else:
            frame = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)

            with st.spinner("Detecting helmets…"):
                result_frame, counts, violations, plate_text, plate_crop = run_detection(
                    frame, models, conf_threshold
                )

            result_rgb = cv2.cvtColor(result_frame, cv2.COLOR_BGR2RGB)
            st.image(result_rgb, use_container_width=True)

            # Stats
            total = counts['helmet'] + counts['no_helmet']
            st.markdown(f"""
<div class="sc">
  <div class="sc-item sc-blue">
    <div class="sc-num">{total}</div><div class="sc-lbl">Riders</div>
  </div>
  <div class="sc-item sc-green">
    <div class="sc-num">{counts['helmet']}</div><div class="sc-lbl">With Helmet</div>
  </div>
  <div class="sc-item sc-red">
    <div class="sc-num">{counts['no_helmet']}</div><div class="sc-lbl">No Helmet</div>
  </div>
</div>""", unsafe_allow_html=True)

            # Plate + violations
            if violations:
                st.markdown('<p class="sec-label">Number Plate — Violation Rider</p>',
                              unsafe_allow_html=True)
                if plate_text and plate_text not in ('NOT DETECTED','N/A'):
                    st.markdown(f'<div class="plate-ok">{plate_text}</div>',
                                  unsafe_allow_html=True)
                else:
                    st.markdown('<div class="plate-na">Plate not detected</div>',
                                  unsafe_allow_html=True)

                st.markdown('<p class="sec-label">Violations</p>', unsafe_allow_html=True)
                tags = "".join([
                    f'<span class="vtag">⚠ Rider #{i+1} — No Helmet</span>'
                    for i in range(counts['no_helmet'])
                ])
                st.markdown(f'<div>{tags}</div>', unsafe_allow_html=True)
            else:
                st.success("✅  All riders wearing helmets. No violations detected.")

            st.markdown("")

            # Downloads
            d1, d2, d3 = st.columns(3)

            img_buf = io.BytesIO()
            Image.fromarray(result_rgb).save(img_buf, format="JPEG", quality=93)
            with d1:
                st.download_button("⬇ Annotated Image", data=img_buf.getvalue(),
                                    file_name="detection_result.jpg", mime="image/jpeg",
                                    use_container_width=True)

            with d2:
                if violations and plate_crop is not None:
                    pil_plate = Image.fromarray(cv2.cvtColor(plate_crop, cv2.COLOR_BGR2RGB))
                    if pil_plate.width < 400:
                        scale = 400/pil_plate.width
                        pil_plate = pil_plate.resize(
                            (int(pil_plate.width*scale), int(pil_plate.height*scale)),
                            Image.LANCZOS)
                    plate_buf = io.BytesIO()
                    pil_plate.save(plate_buf, format="PNG")
                    st.download_button("🔍 Plate Image", data=plate_buf.getvalue(),
                                        file_name="number_plate.png", mime="image/png",
                                        use_container_width=True)
                else:
                    st.button("🔍 No Plate Found", disabled=True, use_container_width=True)

            with d3:
                with st.spinner("Building report…"):
                    xlsx_buf = build_excel(uploaded.name, counts,
                                            violations, plate_text, plate_crop)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button("📊 Excel Report", data=xlsx_buf.getvalue(),
                                    file_name=f"violation_report_{ts}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True)

    else:
        st.markdown("""
<div class="empty-state">
    <div style="font-size:2.6rem">🎯</div>
    <div style="font-weight:600;color:#64748b;margin:8px 0 4px">Results appear here</div>
    <div style="font-size:0.82rem;color:#94a3b8">Upload an image and click Run Detection</div>
</div>""", unsafe_allow_html=True)